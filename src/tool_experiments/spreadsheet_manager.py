from pathlib import Path
from openpyxl import load_workbook, Workbook
from typing import Optional, List, Any
import pandas as pd

class SpreadsheetManager:
    """Manages reading and writing operations for Excel spreadsheets."""

    @classmethod
    def CreateNew(cls, file_path: Path, sheet_name: str = "Sheet1"):
        """Create a new Excel spreadsheet file.
        
        Args:
            file_path: Path where the new spreadsheet should be created
            sheet_name: Name of the first sheet (default: "Sheet1")
            
        Returns:
            SpreadsheetManager instance with the new workbook
        """
        workbook = cls._create_new_workbook(sheet_name)
        cls._save_and_close_workbook(workbook, file_path)
        return cls._create_manager_instance(file_path)

    @classmethod
    def _create_new_workbook(cls, sheet_name: str):
        """Create a new workbook with the specified sheet name."""
        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)
        workbook.create_sheet(title=sheet_name)
        return workbook

    @classmethod
    def _save_and_close_workbook(cls, workbook, file_path: Path):
        """Save and close the workbook."""
        workbook.save(file_path)
        workbook.close()

    @classmethod
    def _create_manager_instance(cls, file_path: Path):
        """Create and return a SpreadsheetManager instance with the new workbook."""
        manager = cls(file_path)
        manager.workbook = load_workbook(file_path)
        manager.is_open = True
        return manager

    def __init__(self, file_path: Path):
        """Initialize with path to spreadsheet file."""
        self.file_path = Path(file_path)
        self.workbook = None
        self.is_open = False

    def open(self):
        """Open the spreadsheet file."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Spreadsheet not found: {self.file_path}")
        self.workbook = load_workbook(self.file_path)
        self.is_open = True

    def close(self) -> None:
        """Close the spreadsheet."""
        if self.workbook:
            self.workbook.close()
        self.workbook = None
        self.is_open = False
    
    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        return self.workbook.sheetnames
    
    def read_cell(self, sheet_name: str, cell_address: str) -> Any:
        """Read value from a specific cell using A1 notation."""
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        sheet = self.workbook[sheet_name]
        return sheet[cell_address].value
    
    def read_cell_coords(self, sheet_name: str, row: int, col: int) -> Any:
        """Read value from a specific cell using row/column coordinates."""
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row, column=col).value
    
    def find_next_empty_cell_in_column(self, sheet_name: str, column: str, start_row: int = 1) -> int:
        """Find the next empty cell in a column starting from start_row.
        
        Args:
            sheet_name: Name of the sheet to search
            column: Column letter (e.g., 'A', 'B', etc.)
            start_row: Row number to start searching from (default: 1)
            
        Returns:
            Row number of the next empty cell
        """
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        sheet = self.workbook[sheet_name]
        col_letter = column.upper()
        max_row = sheet.max_row
        return self._find_empty_cell_in_range(sheet, col_letter, start_row, max_row)

    def _find_empty_cell_in_range(self, sheet, col_letter: str, start_row: int, max_row: int) -> int:
        """Find the next empty cell in the specified column range."""
        for row in range(start_row, max_row + 1):
            cell_value = sheet[f"{col_letter}{row}"].value
            if self._is_cell_empty(cell_value):
                return row
        return max_row + 1

    def _is_cell_empty(self, cell_value: Any) -> bool:
        """Check if a cell value is considered empty."""
        return cell_value is None or str(cell_value).strip() == ""
    
    def readRangeAsDataFrame(self, sheet_name: Optional[str] = None, 
                           start_row: Optional[int] = None, end_row: Optional[int] = None,
                           start_col: Optional[str] = None, end_col: Optional[str] = None) -> pd.DataFrame:
        """Read a range of cells as a pandas DataFrame.
        
        Args:
            sheet_name: Name of the sheet to read (default: first sheet)
            start_row: Starting row number (default: 1)
            end_row: Ending row number (default: auto-detect)
            start_col: Starting column letter (default: 'A')
            end_col: Ending column letter (default: auto-detect)
            
        Returns:
            pandas DataFrame containing the range data
        """
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        sheet_name = self._get_sheet_name_or_default(sheet_name)
        sheet = self.workbook[sheet_name]
        
        start_row, start_col = self._set_default_range_values(start_row, start_col)
        end_col = self._determine_end_column(sheet, start_row, end_col)
        end_row = self._determine_end_row(sheet, start_col, end_row)
        
        start_col_idx, end_col_idx = self._convert_columns_to_indices(start_col, end_col)
        data = self._read_cell_range(sheet, start_row, end_row, start_col_idx, end_col_idx)
        
        return self._create_dataframe_with_headers(data)
    
    def _ensure_workbook_open(self):
        """Ensure the workbook is open before performing operations."""
        if not self.is_open:
            raise RuntimeError("Spreadsheet must be opened first")

    def _get_sheet_name_or_default(self, sheet_name: Optional[str]) -> str:
        """Get the sheet name or use the first sheet if not specified."""
        if sheet_name is None:
            if self.workbook is None:
                raise RuntimeError("Workbook is not open")
            sheet_name = self.workbook.sheetnames[0]
        return sheet_name

    def _set_default_range_values(self, start_row: Optional[int], start_col: Optional[str]) -> tuple[int, str]:
        """Set default values for start row and column."""
        return start_row or 1, start_col or 'A'

    def _determine_end_column(self, sheet, start_row: int, end_col: Optional[str]) -> str:
        """Determine the end column, auto-detecting if not specified."""
        if end_col is None:
            end_col = self._find_last_non_empty_column(sheet, start_row)
        return end_col

    def _determine_end_row(self, sheet, start_col: str, end_row: Optional[int]) -> int:
        """Determine the end row, auto-detecting if not specified."""
        if end_row is None:
            end_row = self._find_last_non_empty_row(sheet, start_col)
        return end_row

    def _convert_columns_to_indices(self, start_col: str, end_col: str) -> tuple[int, int]:
        """Convert column letters to column indices."""
        start_col_idx = self._column_letter_to_index(start_col)
        end_col_idx = self._column_letter_to_index(end_col)
        return start_col_idx, end_col_idx

    def _read_cell_range(self, sheet, start_row: int, end_row: int, start_col_idx: int, end_col_idx: int) -> list:
        """Read the specified cell range and return as a list of lists."""
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col_idx, end_col_idx + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        return data

    def _create_dataframe_with_headers(self, data: list) -> pd.DataFrame:
        """Create a DataFrame from data and set headers if appropriate."""
        df = pd.DataFrame(data)
        if len(data) > 0 and self._should_use_first_row_as_headers(data[0]):
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        return df

    def _should_use_first_row_as_headers(self, first_row: list) -> bool:
        """Determine if the first row should be used as column headers."""
        return any(cell is not None and str(cell).strip() != "" for cell in first_row)
    
    def write_dataframe(self, df: pd.DataFrame, sheet_name: str, start_cell: str = "A1", 
                       include_headers: bool = True) -> None:
        """Write a pandas DataFrame to a specified sheet and location.
        
        Args:
            df: DataFrame to write
            sheet_name: Name of the sheet to write to
            start_cell: Starting cell address (e.g., "A1", "B3")
            include_headers: Whether to include DataFrame column headers (default: True)
        """
        self._ensure_workbook_open()
        if self.workbook is None:
            raise RuntimeError("Workbook is not open")
        sheet = self.workbook[sheet_name]
        start_row, start_col = self._parse_cell_address(start_cell)
        data_start_row = self._write_headers_if_requested(sheet, df, start_row, start_col, include_headers)
        self._write_data_rows(sheet, df, data_start_row, start_col)
        self._save_workbook()
    
    def _write_headers_if_requested(self, sheet, df: pd.DataFrame, start_row: int, start_col: int, include_headers: bool) -> int:
        """Write headers if requested and return the starting row for data."""
        if include_headers:
            for col_idx, header in enumerate(df.columns):
                cell_address = self._get_cell_address(start_row, start_col + col_idx)
                sheet[cell_address] = header
            return start_row + 1
        return start_row

    def _write_data_rows(self, sheet, df: pd.DataFrame, data_start_row: int, start_col: int):
        """Write all data rows to the sheet."""
        for row_idx, row_data in enumerate(df.values):
            for col_idx, value in enumerate(row_data):
                cell_address = self._get_cell_address(data_start_row + row_idx, start_col + col_idx)
                sheet[cell_address] = value

    def _save_workbook(self):
        """Save the workbook to persist changes."""
        if self.workbook:
            self.workbook.save(self.file_path)
    
    def _parse_cell_address(self, cell_address: str) -> tuple[int, int]:
        """Parse cell address (e.g., 'A1') to row and column indices."""
        import re
        
        match = self._match_cell_address_pattern(cell_address)
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_idx = self._column_letter_to_index(col_letter)
        return row_num, col_idx

    def _match_cell_address_pattern(self, cell_address: str):
        """Match cell address against the expected pattern."""
        import re
        match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
        if not match:
            raise ValueError(f"Invalid cell address: {cell_address}")
        return match
    
    def _get_cell_address(self, row: int, col: int) -> str:
        """Convert row and column indices to cell address."""
        col_letter = self._column_index_to_letter(col)
        return f"{col_letter}{row}"
    
    def _find_last_non_empty_column(self, sheet, row: int) -> str:
        """Find the last non-empty column in a given row."""
        max_col = sheet.max_column
        for col in range(max_col, 0, -1):
            cell_value = sheet.cell(row=row, column=col).value
            if not self._is_cell_empty(cell_value):
                return self._column_index_to_letter(col)
        return 'A'
    
    def _find_last_non_empty_row(self, sheet, column: str) -> int:
        """Find the last non-empty row in a given column."""
        max_row = sheet.max_row
        col_idx = self._column_letter_to_index(column)
        
        for row in range(max_row, 0, -1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            if not self._is_cell_empty(cell_value):
                return row
        return 1
    
    def _column_letter_to_index(self, column: str) -> int:
        """Convert column letter to index (A=1, B=2, etc.)."""
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def _column_index_to_letter(self, index: int) -> str:
        """Convert column index to letter (1=A, 2=B, etc.)."""
        result = ""
        while index > 0:
            index -= 1
            result = chr(ord('A') + (index % 26)) + result
            index //= 26
        return result
    
    def __enter__(self):
        """Context manager entry."""
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()